"""
app/routes/reports.py  –  FIX #3: Excel export route (was missing entirely)
"""
import io
import calendar
from datetime import date, timedelta
from flask import Blueprint, send_file
from flask_jwt_extended import jwt_required
from sqlalchemy import func, extract

from app import db
from app.models import Booking, Payment, Guest
from app.utils  import error

reports_bp = Blueprint("reports", __name__)


# ── GET /api/reports/export-monthly-excel  ──────────────────────────────── ADMIN
@reports_bp.get("/export-monthly-excel")
@jwt_required()
def export_monthly_excel():
    """
    Export the last 12 months of booking + revenue data as an Excel file.
    Requires openpyxl (add to requirements.txt).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return error("openpyxl is not installed. Run: pip install openpyxl", 500)

    today    = date.today()
    year_ago = today - timedelta(days=365)

    # ── Fetch booking stats ───────────────────────────────────────────────
    booking_rows = (
        db.session.query(
            extract("year",  Booking.created_at).label("yr"),
            extract("month", Booking.created_at).label("mo"),
            func.count(Booking.id).label("bookings"),
            func.avg(Booking.adults + Booking.youth).label("avg_pax"),
        )
        .filter(Booking.created_at >= year_ago)
        .group_by("yr", "mo")
        .order_by("yr", "mo")
        .all()
    )

    revenue_rows = (
        db.session.query(
            extract("year",  Payment.paid_at).label("yr"),
            extract("month", Payment.paid_at).label("mo"),
            func.sum(Payment.amount).label("rev"),
        )
        .filter(Payment.paid_at >= year_ago)
        .group_by("yr", "mo")
        .all()
    )
    rev_map = {(int(r.yr), int(r.mo)): float(r.rev) for r in revenue_rows}

    # ── Build workbook ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Summary"

    # Header style
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    headers = ["Month", "Total Bookings", "Revenue (₱)", "Avg Pax"]
    col_widths = [20, 18, 20, 12]

    ws.append(["Palacio Feliz — Monthly Report"])
    ws.append([f"Generated: {today.strftime('%B %d, %Y')}"])
    ws.append([])
    ws.append(headers)

    # Style header row
    header_row = 4
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Style title rows
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"].font = Font(italic=True, color="595959", size=10)

    # Data rows
    row_num = header_row + 1
    alt_fill = PatternFill("solid", fgColor="D6E4F0")

    for r in booking_rows:
        yr, mo = int(r.yr), int(r.mo)
        revenue = rev_map.get((yr, mo), 0.0)
        row_data = [
            f"{calendar.month_name[mo]} {yr}",
            int(r.bookings),
            round(revenue, 2),
            round(float(r.avg_pax or 0), 1),
        ]
        ws.append(row_data)

        # Alternate row shading
        fill = alt_fill if (row_num % 2 == 0) else None
        for col_idx in range(1, 5):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border    = thin_border
            cell.alignment = center
            if fill:
                cell.fill = fill
            if col_idx == 3:
                cell.number_format = '#,##0.00'

        row_num += 1

    # Totals row
    total_revenue  = sum(rev_map.values())
    total_bookings = sum(int(r.bookings) for r in booking_rows)
    ws.append(["TOTAL", total_bookings, round(total_revenue, 2), ""])
    total_row = row_num
    for col_idx in range(1, 5):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill("solid", fgColor="BDD7EE")
        cell.border    = thin_border
        cell.alignment = center
        if col_idx == 3:
            cell.number_format = '#,##0.00'

    # ── Stream as file download ───────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"palacio_monthly_report_{today.strftime('%Y_%m')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
